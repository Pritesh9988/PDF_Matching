import openpyxl
from openpyxl.styles import PatternFill

def highlight_cells_with_exceptions(file_path, sheet_name, column_letter, exception_values):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    column = sheet[column_letter]

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for cell in column:
        if cell.value not in exception_values:
            cell.fill = red_fill

    wb.save(file_path)
    wb.close()

# Example usage
file_path = 'example.xlsx'  # Replace with your file path
sheet_name = 'Sheet1'  # Replace with your sheet name
column_letter = 'A'  # Replace with your column letter
exception_values = [1, 3, 5]  # Replace with your exception values

highlight_cells_with_exceptions(file_path, sheet_name, column_letter, exception_values)
